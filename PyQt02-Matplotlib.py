###############################################################################################
# Qtdesigner를 이용한 Python GUI 프로그래밍 교육자료
# Matplotlib을 PyQt에 내장하는 방법
# 엑셀파일을 불러와서 도표에 나타내는 방법 등
###############################################################################################

import openpyxl
import numpy as np

import sys
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5 import uic               # Qtdesigner(XML파일)을 파이썬 모듈에 포함

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure


class winApp(QMainWindow):  # QDialog / QMainWindow
    def __init__(self):
        super().__init__()
        self.ui = uic.loadUi('PyQt02-Matplotlib2.ui', self)

        # Matplotlib 그래프를 위젯으로 만들기
        self.create_mlp_widget()

        # signal/slot 정의
        self.ui.btn.clicked.connect(self.open_xls)
        self.ui.cb.currentIndexChanged.connect(self.cb_function)

    def create_mlp_widget(self):
        # Create the mpl Figure and FigCanvas objects.
        # 6x2.5 inches, 100 dots-per-inch
        fig = Figure((6, 2.5), dpi=100)                     #fig = Figure(): size 지정
        canvas = FigureCanvas(fig)
        canvas.setParent(self.ui.mpl)
        # FigureCanvas.setSizePolicy(self, QSizePolicy.Expanding, QSizePolicy.Expanding)
        toolbar = NavigationToolbar(canvas,self)
        self.ui.tbar_layer.addWidget(toolbar)              # 원하는 위치에

        # 그래프 그리기
        y = (10, 20, 30, 40, 50)
        x = np.arange(5)
        width = 0.35
        axes = fig.add_subplot(111)
        axes.bar(x, y, width)
        axes.set_xticks(x + width/20)
        axes.set_xticklabels(['A','B','C','D','E'])
        axes.grid('on')

        canvas.draw()
        canvas.show()

    def open_xls(self):
        # get the xls filename by file dialog
        fname = QFileDialog.getOpenFileName(self,
            '엑셀파일 열기', '', 'Excel Files (*.xls, *.xlsx)')

        if not fname[0]:
            QMessageBox.about(self, '파일열기', '파일지정이 잘못되었습니다.')
            return

        wb = openpyxl.load_workbook(fname[0])   # Open the Excel file(Workbook)
        ws = wb.active                          # Open the worksheet

        # Set the row and column of table
        self.ui.table.setColumnCount(ws.max_column)
        self.ui.table.setRowCount(ws.max_row-1)

        # Set the column header name
        # header = ws[1]
        # self.ui.table.setHorizontalHeaderLabels(header.value)

        # Create list which stores the results
        dbom = []
        for row in range(ws.max_row):
            line = []
            for col in range(ws.max_column):
                if row == 0:  # header data
                    item = QTableWidgetItem(ws.cell(row+1, col+1).value)
                    item.setBackground(Qt.green)
                    item.setTextAlignment(Qt.AlignLeft)
                    self.ui.table.setHorizontalHeaderItem(col, item)
                else:
                    line.append(ws.cell(row+1, col+1).value)  # Model data 저장
                    if col == 4:
                        item = QTableWidgetItem(str(ws.cell(row+1, col+1).value))
                        item.setTextAlignment(Qt.AlignLeft)
                    elif col == 6 or col == 7:
                        cell_str = '%.2f' % (ws.cell(row+1, col+1).value)
                        item = QTableWidgetItem(cell_str)
                        item.setTextAlignment(Qt.AlignRight)
                    elif col == 8:
                        cell_str = format(ws.cell(row+1, col+1).value, ',')
                        item = QTableWidgetItem(cell_str)
                        item.setTextAlignment(Qt.AlignRight)
                    elif col == 13:
                        if ws.cell(row+1, col+1).value is None:
                            item = QTableWidgetItem(' ')
                    else:
                        item = QTableWidgetItem(str(ws.cell(row+1, col+1).value))
                        item.setTextAlignment(Qt.AlignCenter)

                    self.ui.table.setItem(row-1, col, item)

            if row != 0:
                dbom.append(line)

        # Table option
        self.ui.table.resizeColumnsToContents()  # Column width(auto)
        # self.ui.table.setColumnWidth(4, 200)    # Column width(manual)
        self.ui.table.resizeRowsToContents()  # Row height(auto)
        # self.ui.table.verticalHeader().setVisible(False) # Hide row header
        # self.ui.table.horizontalHeader().setVisible(False) # Hide column header
        self.ui.table.setSortingEnabled(True)  # Sorting
        self.ui.table.setShowGrid(True)  # Grid on
        self.ui.table.setAlternatingRowColors(True)  # Striped table

        # self.ui.table.setColumnHidden(2, True)    # Hide the Column
        # self.ui.table.setRowHidden(0, True)       # Hide the Row

        # # Set the background of the table
        # palette = QPalette()    # Background color
        # # palette.setColor(QPalette.Base, Qt.yellow)          # Yellow
        # # palette.setColor(QPalette.Base, QColor(0,255,0))    # Green
        # palette.setBrush(QPalette.Base, QBrush(QPixmap('lenna_full.jpg')))    # image
        # self.ui.tableWidget.setPalette(palette)

        # Set the background of the cell
        # target_cell = self.ui.tableWidget.item(0, 4)            # Read the cell value
        # target_cell.setBackground(QBrush(Qt.yellow))            # cell Background - Color
        # # target_cell.setBackground(QBrush(QPixmap('.\icons\win_icon.png')))    # cell Background - image
        # target_cell.setForeground(QBrush(QColor(255, 0, 0)))    # text color
        # # target_cell.setFont(QFont('SansSerif', 10))
        # target_cell.setFont(QFont('Helvetica', 8, QFont.Normal, italic=False))
        # # target_cell.setFont(QFont('Times', 12, QFont.Bold, italic=True)) # Font
        # self.ui.tableWidget.currentCellChanged.connect(self.tcell_change)

        self.ui.lbl.setText(fname[0])
        wb.close()
        return dbom


    def cb_function(self):
        index = self.ui.cb.currentIndex()
        msgstr = self.ui.cb.itemText(index)
        self.ui.statusBar().showMessage(str(index) + msgstr)
        # msgstr = self.ui.cb.currentText()


if __name__ == '__main__':
    app = QApplication(sys.argv)    # sys.argv(입력인자) -> PyQt Application 객체 생성
    app.setStyle(QStyleFactory.create('Fusion'))    # header color 변경에 필요함
    win = winApp()                  # my GUI 객체 생성
    win.show()                      # my GUI 보이기
    sys.exit(app.exec_())           # Application 실행 : 무한루프로 이벤트 처리 -> 출력인자: sys.exit

################################################################################################
# 폰트 설정 방법 1
################################################################################################
"""
import matplotlib.pyplot as plt

# matplotlib 폰트설정
# plt.rc('font', family='NanumGothicOTF') # For MacOS
plt.rc('font', family='NanumGothic') # For Windows
print(plt.rcParams['font.family'])

# 폰트 설정 방법 2
import matplotlib
import matplotlib.font_manager as fm
fm.get_fontconfig_fonts()
# font_location = '/usr/share/fonts/truetype/nanum/NanumGothicOTF.ttf'
font_location = 'C:/Windows/Fonts/NanumGothic.ttf' # For Windows
font_name = fm.FontProperties(fname=font_location).get_name()
matplotlib.rc('font', family=font_name)
"""