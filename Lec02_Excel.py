############################################################
# 파일의 경로 다루기
############################################################
"""
import os

cwd = os.getcwd()                   # get current working directory
flist = os.listdir(cwd)             # 현재폴더의 폴더 및 파일 리스트
'Lec02_Excel.py' in flist           # True/False
os.path.exists('Lec02_Excel.py')    # 파일/폴더의 존재여부 확인

os.chdir('..')                      # change directory
os.chdir('Lecture')

fullfname = os.path.join(cwd,'Lec02_Excel.py')  # OS에 맞춰 파일경로를 만들어 줌

os.path.dirname(fullfname)                      # Directory만 얻기
os.path.basename(fullfname)                     # Filename만 얻기
path, fname = os.path.split(fullfname)          # Directory, Filename 각각 얻기

BASE_DIR = os.path.dirname(os.path.abspath(__file__))   # 현재파일의 경로 구하기

############################################################
# 텍스트 파일 불러오기
############################################################
import os

# 파일경로 설정 (D:\\12-Python\\Lecture\\hello.txt)
fname = os.path.join(os.getcwd(), 'hello.txt')
fname = 'D:\\12-Python\\Lecture\\hello.txt'
fname = r'D:\12-Python\Lecture\hello.txt'
# fname = os.path.join(os.getcwd(), '인사말.txt')


# 파일객체 생성: r,w,a/t,b
f = open(fname, 'rt', encoding='utf-8')
f.encoding      # 기본 인코딩방식 : 'cp949'

# 파일 읽기
contents = f.read()             # 방법1 : 한꺼번에 읽기 - read(n) : n문자
print(contents)

lines = f.readlines()           # 방법2 : 모든 라인을 리스트로 반환
print(lines)
f.seek(0)                       # 파일포인터 이동 (byte 단위)
f.tell()                        # 파일포인터 위치 반환
for line in lines:              # 한 라인씩 순회하며 읽음
    print(line, end='')

while True:
    line = f.readline()         # 방법3 : 한 라인씩 읽기
    if not line:
        break
    print(line, end='')

# 파일객체를 닫음
f.close()

############################################################
# 텍스트 파일 저장하기
############################################################
import os


# 파일 출력
f = open('D:\\12-Python\\Lecture\\filewrite.txt', 'wt', encoding='utf-8')

f.write('write 사용\n첫째 줄\n둘째 줄\n셋째 줄\n')     # 글자수 반환: 24
f.flush()

lines = ['writelines 사용\n', '첫째 줄\n', '둘째 줄\n', '셋째 줄\n']
f.writelines(lines)
f.close()

# 파일 컨텍스트를 이용한 편의문법 : 코딩이 줄고, 자동 close
fname = os.path.join(os.getcwd(), '인사말.txt')
with open(fname, 'r', encoding='utf-8') as f:
    contents = f.read()
    print(contents)


############################################################
# Openpyxl vs. MS COM 모듈
############################################################
import os
import time
import openpyxl
import win32com.client


def load_xls_by_openpyxl(fname):
    openpyxl.load_workbook(fname)


def load_xls_by_mscom(fname):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.Workbooks.Open(fname)
    excel.Quit()


if __name__ == '__main__':
    fname = os.path.join(os.getcwd(), 'bom_whole.xlsx')

    # OpenPyXl 불러오기 시간
    t0 = time.time()
    for i in range(5):
        load_xls_by_openpyxl(fname)  # 27초/회 소요
    print('OpenPyXl 불러오기 : {:.1f} 초'.format((time.time()-t0)/5))

    # Win2Com 불러오기 시간
    t0 = time.time()
    for i in range(5):
        load_xls_by_mscom(fname)       # 7초/회 소요, full path filename
    print('Win2Com 불러오기 : {:.1f} 초'.format((time.time()-t0)/5))


############################################################
# Openpyxl 모듈(1) - 기존문서를 불러와서 작업하고 저장하고 닫기
############################################################
import openpyxl

xls = openpyxl.load_workbook('재고장.xlsx')     # Workbook 열기

xls.sheetnames       # 시트목록을 리스트로 반환 ['물류재고장', '입출고내역', '랙번호']
sheet = xls.active   # 현재 활성화된 sheet를 가져옴
sheet.title          # '물류재고장'
sheet = xls['랙번호'] # '랙번호' sheet 가져오기
sheet.max_row        # 503
sheet.max_column     # 1

sheet['A2']             # <Cell '랙번호'.A2>
sheet['A2'].value       # '0-1-1'
sheet['A2'].coordinate  # 'A2'

sheet.cell(row=2, column=1)     # 숫자 인덱스를 통해 루프를 돌릴 수 있다
sheet.cell(2, 1).value
sheet.cell(2, 1).coordinate
for i in range(1, 8, 2):
    print(i, sheet.cell(i, 1).value)
    # 1 랙번호
    # 3 0-1-2
    # 5 0-2-1
    # 7 0-2-3


# 셀값 가져오기
sheet = xls['물류재고장']
area = sheet['A1:C3']
# ((<Cell '물류재고장'.A1>, <Cell '물류재고장'.B1>, <Cell '물류재고장'.C1>),
#  (<Cell '물류재고장'.A2>, <Cell '물류재고장'.B2>, <Cell '물류재고장'.C2>),
#  (<Cell '물류재고장'.A3>, <Cell '물류재고장'.B3>, <Cell '물류재고장'.C3>))
area[0][1]              # <Cell '물류재고장'.B1>
for row in area:
    for cell in row:
        print(cell.value, end='\t\t')
    print('')

row = sheet[1]
str_row = []
for cell in row:
    str_row.append(cell.value)      # Print, print -> 줄바꿈 (end='\n'), print(, ) -> 공백 (sep=',')
print(", ".join(str_row))

rows = sheet[2:5]
for row in rows:
    for cell in row:
        print(cell.value, end=' / ')
    print('')

col = sheet['A']
for cell in col:
    print(cell.value)

cols = sheet['A:B']
trans_cols = zip(*cols)     # zip + unpacking --> 행과 열의 transpose
for row in trans_cols:
    for cell in row:
        print(cell.value, end='\t')
    print('')

xls.close()

# 리스트의 행과 열을 바꾸기 : zip + unpacking
A = [[1, 2, 3], [4, 5, 6]]
list(zip(*A))           # [(1, 4), (2, 5), (3, 6)]

############################################################
# Openpyxl 모듈(2) - 새 문서 만들고 저장하기
############################################################
from openpyxl import Workbook

xls = Workbook()                    # 새로운 엑셀파일 생성

xls.sheetnames                      # default로 'Sheet' 시트 생성
sheet = xls.active                  # 현재 활성시트
sheet.title = 'Stock'               # sheetname 변경

xls.create_sheet('가나다')           # 마지막 위치에 '가나타' 시트 추가
xls.create_sheet('시트2', 1)         # 해당위치 뒤에 '시트2' 시트 추가

sheet['A1'].value = 'Rack'
# sheet['A1'] = 'Rack'             # value 생략 가능
sheet.cell(2,1).value = '0-0-1'    # 인덱스가 (1,1)부터 가능, value 생략불가

sheet.merge_cells('B1:C1')          # 셀 합치기
sheet['B1'].value = 'B1과 C1 결합'
sheet.merge_cells('A4:C6')
sheet['A4'].value = 'A4:C6 범위 합치기'

sheet['D6'].value = '=SUM(D1:D5)'

xls.save('재고장_new.xlsx')          # 저장
xls.close()                         # 닫기


############################################################
# Openpyxl 모듈(3) - racklist 만들기
############################################################
from openpyxl import load_workbook
from datetime import timedelta, datetime


def find_inventory_by_item(itemcode):
    '''재고량이 1,000보다 적은 데이터를 반환'''
    xls = load_workbook('재고장.xlsx')
    sheet = xls['물류재고장']

    data = []
    for row in sheet[2:sheet.max_row]:
        if row[3].value == itemcode:
            data.append(row)

    xls.close()
    return data


def find_inventory_by_date(date):
    '''생산일이 오래된 데이터 검색'''
    xls = load_workbook('재고장.xlsx')
    sheet = xls['물류재고장']

    data = []
    for row in sheet[2:sheet.max_row]:
        if row[7].value < date:
            data.append(row)

    xls.close()
    return data


if __name__ == '__main__':
    # inventory = find_inventory_by_item('SV0185')
    # today = datetime(2018, 5, 1)
    today = datetime.today()                    # datetime
    ref_day = today - timedelta(days=365*3)     # timedelta
    str_date = ref_day.strftime('%Y-%m-%d')     # strftime
    inventory = find_inventory_by_date(str_date)
    for row in inventory:
        print(row[0].value, end='\t')               # 랙번호
        print(row[3].value, end='\t')               # 품목코드
        print(row[4].value, end='\t')               # 품목명
        print(row[5].value, end='\t')
        print(format(row[6].value, ','), end='\t')
        print(row[7].value, end='\n')

############################################################
# Openpyxl 모듈(4) - 셀서식 지정하기
############################################################
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# 상수
TITLE_CELL_COLOR = "AA8866"

wb = openpyxl.load_workbook('재고장.xlsx')
ws = wb['물류재고장']

# ws.append(['1-2-3', 'SI1234', '추가된 사출성형 품목', '사출', 6000, '2021-03-29'])

# 틀 고정
ws.freeze_panes = "C2"

# 열 크기 지정
col_widths = {'A':8, 'B':13, 'C':50, 'D':10, 'E':10, 'F':15}
for pos, width in col_widths.items():
    ws.column_dimensions[pos].width = width

# 행 높이 지정
# ws.row_dimensions[1].height = 30
for i in range(1, ws.max_row+1):
    ws.row_dimensions[i].height = 20    #
    ws.cell(i, 5).number_format = '#,##0'
    if i != 1 and int(ws.cell(i, 5).value) < 1000:
        ws.cell(i, 5).font = Font(bold=True)

# 폰트 지정
font_header = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
for cols in ws['A1':'F1']:
    for cell in cols:
        cell.fill = PatternFill(patternType='solid', fgColor=TITLE_CELL_COLOR)
        cell.alignment = Alignment(horizontal='distributed')
        cell.font = font_header

# 셀 테두리 지정
side = Side(style='thin', color='000000')
border = Border(left=side, right=side, top=side, bottom=side)
for row in ws:
    for cell in row:
        cell.border = border

# 행과 열 숨기기
# ws.column_dimensions['A'].hidden = True     # 열 숨기기
# # ws.column_dimensions['A'].hidden = False    # 숨긴 열 표시하기
# ws.row_dimensions[1].hidden = True

wb.save('재고장_new.xlsx')
wb.close()



############################################################
# Openpyxl 모듈(5) - 간단한 chart 그리기
############################################################
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active

for i in range(10):
    ws.append([i])

values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, "C1")

wb.save("SampleChart.xlsx")
wb.close()

"""
############################################################
# Openpyxl 모듈(6) - 다양한 chart 그리기
############################################################
import openpyxl
from openpyxl.chart import Reference, BarChart, LineChart, AreaChart, PieChart, RadarChart

wb = openpyxl.load_workbook('재고장.xlsx')
ws = wb['물류재고장']

# chart 데이터 참조범위 지정
data1 = Reference(ws, min_col=5, min_row=1, max_col=6, max_row=12)
# data2 = Reference(ws, 5, 2, 5, 12)
labels = Reference(ws, 1, 2, 1, 12)

# 차트 종류 지정
# chart = BarChart()    # 막대그래프
# chart.type = 'bar'    # 가로(bar), 세로(col) - 디폴트값
# chart = LineChart()     # 꺾은 선 그래프
# chart = AreaChart()     # 영역형 그래프
# chart = PieChart()      # 원형 차트
chart = RadarChart()    # 방사형 차트

chart.grouping = 'stacked'  # 누적
# chart.overlap = 100       # for barchart: 100


chart.add_data(data1, titles_from_data=True) # title...=True: 첫 셀값을 계열값(범례)로 사용
# chart.add_data(data2)
chart.set_categories(labels)

chart.title = '재고현황'              # 차트 제목
# chart.x_axis.title = '랙번호'        # x축 제목
# chart.y_axis.title = '재고량'        # y축 제목

chart.height = 10                   # 차트 가로 크기
chart.width = 20                    # 차트 세로 크기

ws.add_chart(chart, 'B14')          # chart를 시트의 B14에 삽입

wb.save('재고장_new.xlsx')
wb.close()


"""
############################################################
# COM을 이용한 엑셀파일 다루기(1) - 새 문서 만들고 저장하기
############################################################
import win32com.client

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = True
xls = excel.Workbooks.Add()         # 디폴트로 'Sheet1' 시트가 생성
sheet = xls.ActiveSheet
sheet.Name = 'Sheet'                  # 'Sheet1'
sheet1 = xls.Worksheets.Add()
sheet2 = xls.Worksheets.Add(Before=None, After=sheet)   # sheet 뒤에 추가
sheet3 = xls.Worksheets.Add(Before=sheet)               # sheet 앞에 추가

sheet.Cells(1, 1).Value = "Hello"   # Cells(좌표) - 값 할당
sheet.Range('A3:C4').Value = 1      # Range(범위) - 값 할당
sheet.Range('A2:C2, A5:C5').Value = 2

# AutoFill
sheet.Range('D1:E1').Value = 1
sheet.Range('E2').Value = 2
sheet.Range('D1').AutoFill(sheet.Range('D1:D5'))        # 1 / 1 / 1 / 1 / 1
sheet.Range('E1:E2').AutoFill(sheet.Range('E1:E5'))     # 1 / 2 / 3 / 4 / 5

# 목사 및 붙여넣기
sheet.Range('A5:E5').Copy(sheet.Range('A7:E7'))
sheet.Range('A5:E5').Copy(sheet2.Range('A1'))

# 행/열/셀 삽입
sheet.Rows(5).Insert()
sheet.Columns(2).Insert()
sheet.Cells(3, 3).Insert()

# 수식 넣기
sheet.Range("A1:J10").Formula = "=row()*column()"

xls.SaveAs('D:\\12-Python\\Lecture\\재고장_com.xlsx')          # 저장
xls.Save()
xls.Close()
excel.Quit()


############################################################
# COM을 이용한 엑셀파일 다루기(2) - 기존문서를 열고 작업하기
############################################################
import os
import win32com.client

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = True

xls = excel.Workbooks.Open(os.path.join(os.getcwd(), "재고장_com.xlsx"))
xls.Worksheets.Count                # Sheet 수
sheet_names = [sheet.name for sheet in xls.Sheets]
sheet = xls.Worksheets.Add()         # 추가된 시트가 ActiveSheet가 되고, 이전 Active sheet 앞에 추가
sheet.Name = "추가"                   # xls.ActiveSheet.Name -> '추가'
sheet = xls.Worksheets("Sheet1")    # 시트 선택
# sheet = xls.Worksheets(2)         # 2번째 시트 가져오기

# 셀 색상 지정
sheet.Range('A3:C4').Interior.ColorIndex = 27   # 노란색
sheet.Range('B:B').Interior.ColorIndex = 10     # 녹색

# 셀 폭/높기 조정
sheet.Columns(2).ColumnWidth = 10       # 2열을 10글자 폭으로 조정
sheet.Range("C:C").ColumnWidth = 10
sheet.Rows(1).RowHeight = 30            # 첫째 줄을 30으로 조정(디폴트는 16.5)
sheet.Range('2:5').RowHeight = 20

# 정렬
sheet.Range("2:2").VerticalAlignment = win32com.client.constants.xlCenter
sheet.Rows(1).VerticalAlignment = win32com.client.constants.xlCenter

xls.Save()
xls.Close()
excel.quit()


############################################################
# COM을 이용한 엑셀파일 다루기(3) - 실습
############################################################
import win32com.client

excel = win32com.client.Dispatch('Excel.Application')
excel.Visible = True
wb = excel.Workbooks.Add()
ws = wb.Worksheets('Sheet1')
ws.Name = 'Built with Python'
ws.Cells(1, 1).Value = 'Hello Excel'

for i in range(1, 5):
    ws.Cells(2, i).Value = i    # Cell 루프는 파이썬과 엑셀 간의 통신 Overhead로 효율이 저하됨

ws.Range(ws.Cells(3, 1), ws.Cells(3, 4)).Value = [5, 6, 7, 8] # 일괄처리 Range
ws.Range("A4:D4").Value = [i for i in range(9, 13)]
# ws.Cells(5, 4).Value = '=SUM(A2:D4)'
ws.Cells(5, 4).Formula = '=SUM(A2:D4)'
ws.Cells(5, 4).Font.Size = 16
ws.Cells(5, 4).Font.Bold = True
#ws.Cells(5, 4).Font.Name = '맑은고딕'
ws.Cells(5, 4).Font.Color = RGB(255, 0, 0)

ws.Range("A1:A5").VerticalAlignment = 2         # Up(1), Center(2), Down(3)
ws.Range("A1:D1").HorizontalAlignment = 3       # Left(1), Center(3), Right(4)
ws.Range("B1:B5").NumberFormat = "$###,##0.00"  # "0.0%
ws.Columns.AutoFit()


"""
'''
############################################################
# COM을 이용한 엑셀파일 다루기(4) - 엑셀달력 만들기
############################################################
import time
import calendar
import win32com.client
from win32api import RGB    # RGB 색상 지정


week_name = ('일', '월', '화', '수', '목', '금', '토')             # 한글 요일
# week_name = ('SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT') # 영문 요일


def open_xls():
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    wb = excel.Workbooks.Add()
    ws = wb.Worksheets("Sheet1")
    ws.Name = '달력'
    return ws, wb, excel


def format_calendar(ws):
    """달력의 서식을 지정"""
    # 달력의 요일 양식 지정
    ws.Range("A2:G2").Font.Size = 12
    ws.Range("A2:G2").ColumnWidth = 11
    ws.Range("A2:G2").VerticalAlignment = 2     # center
    ws.Range("A2:G2").HorizontalAlignment = 3   # center
    ws.Range("A2:G2").Orientation = 0
    ws.Range("A2:G2").Font.Size = 12
    ws.Range("A2:G2").Font.Bold = True
    ws.Range("A2:G2").RowHeight = 20
    ws.Range("A2:G2").Value = week_name

    # 달력의 날짜 양식 지정
    ws.Range("A3:G8").HorizontalAlignment = 2   # left
    ws.Range("A3:G8").VerticalAlignment = 1     # Up
    ws.Range("A3:G8").Font.Size = 18
    ws.Range("A3:G8").Font.Bold = True
    ws.Range("A3:G8").RowHeight = 50
    ws.Range("A2:A8").Font.Color = RGB(255, 0, 0)   # 일요일(적)
    ws.Range("G2:G8").Font.Color = RGB(0, 0, 255)   # 토요일(청)
    ws.Range("A2:G8").Borders.LineStyle = 12        # 이중선


def display_days(ws, year, month):
    """달력 양식에 숫자(날짜)를 채움"""
    week_num, days = get_month_info(year, month)

    ws.Cells(1, 1).Value = str(year) + '년'
    ws.Cells(1, 4).Value = str(month)
    ws.Cells(1, 4).RowHeight = 40
    ws.Cells(1, 4).VerticalAlignment = 2  # center
    ws.Cells(1, 4).HorizontalAlignment = 3  # center
    ws.Cells(1, 4).Orientation = 0
    ws.Cells(1, 4).Font.Size = 30
    ws.Cells(1, 4).Font.Bold = True

    for i in range(1, days + 1):
        ws.Range("A3:G8")[i + week_num - 1].Value = i


def get_month_info(year, month):
    """해당월의 시작요일, 날수 구하기"""
    month_info = calendar.monthrange(year, month)           # (1일의 요일, 말일) 형태의 튜플을 반환
    wknum = (month_info[0]+1) % 7                           # 첫날의 요일 : 월요일(0) -> 일요일(0)
    num_days = month_info[1]                                # 해당월의 날수
    print('{}년 {}월 1일은 {}요일이고, 말일은 {}일입니다.'
          .format(year, month, week_name[wknum], num_days))
    return wknum, num_days


if __name__ == '__main__':
    year, month = 2021, 3               # 연도, 월 설정

    sheet, xls, excel = open_xls()      # 엑셀파일 열기
    format_calendar(sheet)              # 달력 양식 지정
    display_days(sheet, year, month)    # 달력 내용 지정
    time.sleep(5)

    # xls.SaveAs('calendar.xlsx')       # 엑셀파일 저장/종료
    xls.Saved = 2                       # 저장 대화상자 : 예(1), 아니오(2)
    xls.Close()
    excel.Quit()
    
'''


